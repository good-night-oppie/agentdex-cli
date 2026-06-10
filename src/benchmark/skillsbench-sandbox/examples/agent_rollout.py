#!/usr/bin/env python3
"""Gym-style RL rollout example for SkillsBench sandbox server.

This script demonstrates:
- A Gym-like env wrapper with reset/step/close
- HTTP interaction with the sandbox server endpoints
- A pluggable command agent interface (rule-based and OpenAI)
"""

import argparse
from abc import ABC, abstractmethod
import json
import re
import subprocess
import textwrap
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib import error, request
from urllib.parse import urlparse

try:
    import gymnasium as gym
except ImportError:
    gym = None


_GYM_BASE = gym.Env if gym is not None else object


class SandboxHTTPClient(object):
    def __init__(self, base_url: str, timeout_sec: int = 60):
        self.base_url = base_url.rstrip("/")
        self.timeout_sec = timeout_sec

    def _json_request(self, method: str, path: str, payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        url = self.base_url + path
        body = None
        headers = {"Content-Type": "application/json"}
        if payload is not None:
            body = json.dumps(payload).encode("utf-8")

        req = request.Request(url=url, data=body, headers=headers, method=method)
        try:
            with request.urlopen(req, timeout=self.timeout_sec) as resp:
                raw = resp.read().decode("utf-8")
        except error.HTTPError as exc:
            err_raw = exc.read().decode("utf-8", errors="replace")
            try:
                err_body = json.loads(err_raw)
            except Exception:
                err_body = {"error": err_raw}
            raise RuntimeError("HTTP {} {} failed: {}".format(exc.code, url, err_body))
        except error.URLError as exc:
            hint = ""
            if isinstance(getattr(exc, "reason", None), ConnectionRefusedError):
                hint = (
                    " | Hint: start sandbox server first, e.g. "
                    "`python3 start_sandbox_server.py --config ./sandbox_config.yaml`, "
                    "then check `curl -s http://127.0.0.1:8080/health`."
                )
            raise RuntimeError("Request to {} failed: {}{}".format(url, exc, hint))

        if not raw:
            return {}
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            raise RuntimeError("Non-JSON response from {}: {}".format(url, raw[:500]))

    def create_env(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        return self._json_request("POST", "/envs", payload)

    def get_instruction(self, env_id: str) -> Dict[str, Any]:
        return self._json_request("GET", "/envs/{}/instruction".format(env_id))

    def step(self, env_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        return self._json_request("POST", "/envs/{}/step".format(env_id), payload)

    def evaluate(self, env_id: str, payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        return self._json_request("POST", "/envs/{}/evaluate".format(env_id), payload or {})

    def reset_env(self, env_id: str) -> Dict[str, Any]:
        return self._json_request("POST", "/envs/{}/reset".format(env_id), {})

    def delete_env(self, env_id: str, remove_image: bool = False) -> Dict[str, Any]:
        query = "?remove_image=true" if remove_image else ""
        return self._json_request("DELETE", "/envs/{}{}".format(env_id, query))


class ManagedSandboxServer(object):
    def __init__(self, process: subprocess.Popen, log_path: Path, started_cmd: str):
        self.process = process
        self.log_path = log_path
        self.started_cmd = started_cmd

    def stop(self) -> None:
        if self.process.poll() is not None:
            return
        self.process.terminate()
        try:
            self.process.wait(timeout=8)
        except Exception:
            self.process.kill()
            self.process.wait(timeout=5)


class InteractionLogger(object):
    """Append-only plain text logger that mirrors printed output."""

    def __init__(self, path: Optional[Path]):
        self.path = path.resolve() if path is not None else None
        self._log_file = None  # type: Optional[Any]
        if self.path is not None:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self._log_file = self.path.open("a", encoding="utf-8")

    def log_line(self, message: str) -> None:
        if self._log_file is None:
            return
        self._log_file.write(str(message) + "\n")
        self._log_file.flush()

    def log_event(self, event: str, payload: Optional[Dict[str, Any]] = None) -> None:
        # Backward-compatible shim: only persist human-readable print lines.
        del event
        if not isinstance(payload, dict):
            return
        message = payload.get("message")
        if message is None:
            return
        self.log_line(str(message))

    def close(self) -> None:
        if self._log_file is None:
            return
        self._log_file.close()
        self._log_file = None


def _print_and_log(
    message: str,
    interaction_logger: Optional[InteractionLogger] = None,
    event: Optional[str] = None,
    payload: Optional[Dict[str, Any]] = None,
) -> None:
    del event
    del payload
    print(message)
    if interaction_logger is not None:
        interaction_logger.log_line(message)


def _parse_server_url(base_url: str) -> Tuple[str, int]:
    parsed = urlparse(base_url)
    if parsed.scheme not in ("http", ""):
        raise ValueError("server-url must use http, got: {}".format(base_url))
    host = parsed.hostname or "127.0.0.1"
    port = parsed.port or 80
    return host, port


def _is_server_healthy(base_url: str, timeout_sec: int = 3) -> bool:
    probe = SandboxHTTPClient(base_url, timeout_sec=timeout_sec)
    try:
        result = probe._json_request("GET", "/health")
    except Exception:
        return False
    return str(result.get("status", "")).lower() == "ok"


def _tail_text(path: Path, max_chars: int = 1200) -> str:
    if not path.exists():
        return ""
    text = path.read_text(errors="replace")
    if len(text) <= max_chars:
        return text
    return text[-max_chars:]


def _extract_toml_section(text: str, section: str) -> str:
    pattern = r"(?ms)^\s*\[" + re.escape(section) + r"\]\s*(.*?)(?=^\s*\[|\Z)"
    match = re.search(pattern, text)
    return match.group(1) if match else ""


def _extract_toml_string(text: str, section: str, key: str) -> Optional[str]:
    body = _extract_toml_section(text, section)
    if not body:
        return None
    pattern = r"(?m)^\s*" + re.escape(key) + r"\s*=\s*\"([^\"]+)\""
    match = re.search(pattern, body)
    if not match:
        return None
    return match.group(1).strip()


def _extract_toml_float(text: str, section: str, key: str) -> Optional[int]:
    body = _extract_toml_section(text, section)
    if not body:
        return None
    pattern = r"(?m)^\s*" + re.escape(key) + r"\s*=\s*([0-9]+(?:\.[0-9]+)?)"
    match = re.search(pattern, body)
    if not match:
        return None
    try:
        return int(float(match.group(1)))
    except Exception:
        return None


def _extract_frontmatter_value(text: str, key: str) -> str:
    lines = text.splitlines()
    if not lines or lines[0].strip() != "---":
        return ""
    end_index = -1
    for i in range(1, len(lines)):
        if lines[i].strip() == "---":
            end_index = i
            break
    if end_index < 0:
        return ""

    want = key.strip().lower()
    for i in range(1, end_index):
        line = lines[i]
        if ":" not in line:
            continue
        k, v = line.split(":", 1)
        if k.strip().lower() == want:
            return v.strip().strip('"').strip("'")
    return ""


def _load_skill_metadata(task_dir: Path) -> List[Dict[str, str]]:
    skills_root = task_dir / "environment" / "skills"
    if not skills_root.is_dir():
        return []

    metadata = []
    for skill_dir in sorted(skills_root.iterdir(), key=lambda p: p.name):
        if not skill_dir.is_dir():
            continue
        skill_md = skill_dir / "SKILL.md"
        if not skill_md.exists():
            continue

        text = skill_md.read_text(errors="replace")
        skill_name = _extract_frontmatter_value(text, "name") or skill_dir.name
        description = _extract_frontmatter_value(text, "description") or "No description provided."
        metadata.append(
            {
                "name": skill_name,
                "description": description,
                # OpenAI skills-style: expose location in context so model can decide to load.
                "path": "/root/.codex/skills/{}".format(skill_dir.name),
            }
        )
    return metadata


def _difficulty_to_level(difficulty: Optional[str]) -> str:
    """Map task.toml difficulty to SkillsBench level (Core/Extended/Extreme).

    Paper spec: Core=10 rounds, Extended=30 rounds, Extreme=50 rounds.
    """
    diff = str(difficulty or "").strip().lower()
    if diff in ("core", "easy"):
        return "core"
    if diff in ("extreme", "hard"):
        return "extreme"
    # "medium" and anything else → extended
    return "extended"


def _default_rounds_for_level(level: str) -> int:
    lv = str(level).strip().lower()
    if lv == "core":
        return 10
    if lv == "extreme":
        return 50
    return 30


def _resolve_task_path(
    project_root: Path,
    dataset: Optional[str],
    task_id: Optional[str],
    task_path: Optional[str],
) -> Optional[Path]:
    if task_path:
        direct = Path(task_path).expanduser()
        if direct.exists():
            return direct.resolve()
        nested = project_root / "skillsbench" / task_path
        if nested.exists():
            return nested.resolve()
        return direct.resolve()
    if not dataset or not task_id:
        return None
    candidate = project_root / "skillsbench" / str(dataset) / str(task_id)
    if candidate.exists():
        return candidate.resolve()
    return None


def _load_task_meta(task_dir: Optional[Path]) -> Dict[str, Any]:
    meta = {
        "task_dir": str(task_dir) if task_dir else "",
        "difficulty": "",
        "level": "extended",
        "default_rounds": 30,
        "agent_timeout_sec": None,
        "verifier_timeout_sec": None,
        "has_skills_dir": False,
        "skills_metadata": [],
    }  # type: Dict[str, Any]
    if task_dir is None:
        return meta

    skills_dir = task_dir / "environment" / "skills"
    meta["has_skills_dir"] = skills_dir.exists() and skills_dir.is_dir()
    meta["skills_metadata"] = _load_skill_metadata(task_dir)

    task_toml = task_dir / "task.toml"
    if not task_toml.exists():
        return meta

    text = task_toml.read_text(errors="replace")
    difficulty = _extract_toml_string(text, "metadata", "difficulty")
    level = _difficulty_to_level(difficulty)
    meta["difficulty"] = difficulty or ""
    meta["level"] = level
    meta["default_rounds"] = _default_rounds_for_level(level)
    meta["agent_timeout_sec"] = _extract_toml_float(text, "agent", "timeout_sec")
    meta["verifier_timeout_sec"] = _extract_toml_float(text, "verifier", "timeout_sec")
    return meta


def _start_server_if_needed(
    base_url: str,
    config_path: Optional[str],
    startup_timeout_sec: int,
    log_path: Path,
    interaction_logger: Optional[InteractionLogger] = None,
) -> Optional[ManagedSandboxServer]:
    if _is_server_healthy(base_url):
        _print_and_log("[server] already healthy at {}".format(base_url), interaction_logger)
        return None

    host, port = _parse_server_url(base_url)
    if host not in ("127.0.0.1", "localhost", "::1"):
        raise RuntimeError(
            "auto-start-server only supports local server-url. current host={}".format(host)
        )

    project_root = Path(__file__).resolve().parents[1]
    server_script = project_root / "start_sandbox_server.py"
    if not server_script.exists():
        raise RuntimeError("Cannot find start script: {}".format(server_script))

    resolved_config = None  # type: Optional[Path]
    if config_path:
        resolved_config = Path(config_path).expanduser().resolve()
        if not resolved_config.exists():
            raise RuntimeError("Config file not found: {}".format(resolved_config))
    else:
        default_cfg = project_root / "sandbox_config.yaml"
        if default_cfg.exists():
            resolved_config = default_cfg

    cmd = ["python3", str(server_script)]
    if resolved_config is not None:
        cmd.extend(["--config", str(resolved_config)])
    cmd.extend(["--host", host, "--port", str(port)])

    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_file = log_path.open("a")
    log_file.write("\n[auto-start] {}\n".format(" ".join(cmd)))
    log_file.flush()

    process = subprocess.Popen(
        cmd,
        cwd=str(project_root),
        stdout=log_file,
        stderr=subprocess.STDOUT,
        close_fds=True,
    )

    deadline = time.time() + max(1, startup_timeout_sec)
    while time.time() < deadline:
        if process.poll() is not None:
            log_file.close()
            tail = _tail_text(log_path)
            raise RuntimeError(
                "Auto-started server exited early (code {}). Log tail:\n{}".format(
                    process.returncode,
                    tail,
                )
            )
        if _is_server_healthy(base_url):
            log_file.close()
            _print_and_log(
                "[server] auto-started at {} (pid={})".format(base_url, process.pid),
                interaction_logger,
            )
            return ManagedSandboxServer(
                process=process,
                log_path=log_path,
                started_cmd=" ".join(cmd),
            )
        time.sleep(0.5)

    log_file.close()
    process.terminate()
    raise RuntimeError(
        "Timed out waiting for server health at {}. Check log: {}".format(base_url, log_path)
    )


class SandboxGymEnv(_GYM_BASE):
    """Gym-like wrapper over the sandbox HTTP API.

    reset() returns (observation, info)
    step(action) returns (observation, reward, terminated, truncated, info)
    """

    def __init__(
        self,
        client: SandboxHTTPClient,
        dataset: Optional[str] = None,
        task_id: Optional[str] = None,
        task_path: Optional[str] = None,
        rebuild: bool = False,
        max_episode_steps: int = 12,
        evaluate_every_step: bool = False,
        step_defaults: Optional[Dict[str, Any]] = None,
        eval_defaults: Optional[Dict[str, Any]] = None,
        auto_cleanup: bool = True,
        interaction_logger: Optional[InteractionLogger] = None,
    ):
        self.client = client
        self.dataset = dataset
        self.task_id = task_id
        self.task_path = task_path
        self.rebuild = rebuild
        self.max_episode_steps = max_episode_steps
        self.evaluate_every_step = evaluate_every_step
        self.step_defaults = step_defaults or {}
        self.eval_defaults = eval_defaults or {}
        self.auto_cleanup = auto_cleanup
        self.interaction_logger = interaction_logger

        self.env_id = None  # type: Optional[str]
        self.instruction = ""
        self.episode_step = 0
        self.last_reward = None  # type: Optional[float]
        self.last_step_result = None  # type: Optional[Dict[str, Any]]

    def _task_payload(self) -> Dict[str, Any]:
        payload = {"rebuild": self.rebuild}
        if self.task_path:
            payload["task_path"] = self.task_path
        else:
            payload["dataset"] = self.dataset
            payload["task_id"] = self.task_id
        return payload

    def _make_observation(self) -> Dict[str, Any]:
        step_result = self.last_step_result or {}
        return {
            "instruction": self.instruction,
            "last_command": step_result.get("command", ""),
            "last_stdout": step_result.get("stdout", ""),
            "last_stderr": step_result.get("stderr", ""),
            "last_exit_code": step_result.get("exit_code"),
            "last_reward": self.last_reward,
            "step_count": self.episode_step,
        }

    def reset(self, seed: Optional[int] = None, options: Optional[Dict[str, Any]] = None) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        del seed
        options = options or {}
        force_new = bool(options.get("new_env", False))

        if self.env_id is None or force_new:
            if force_new and self.env_id and self.auto_cleanup:
                try:
                    self.client.delete_env(self.env_id)
                except Exception:
                    pass
            create_resp = self.client.create_env(self._task_payload())
            self.env_id = create_resp["env_id"]
            info = {"env_id": self.env_id, "create_response": create_resp}
        else:
            reset_resp = self.client.reset_env(self.env_id)
            info = {"env_id": self.env_id, "reset_response": reset_resp}

        instruction_resp = self.client.get_instruction(self.env_id)
        self.instruction = str(instruction_resp.get("instruction", ""))

        if self.interaction_logger is not None:
            self.interaction_logger.log_event(
                "sandbox.reset",
                {
                    "env_id": self.env_id,
                    "force_new": force_new,
                    "response": info,
                },
            )
            self.interaction_logger.log_event(
                "sandbox.instruction",
                {
                    "env_id": self.env_id,
                    "instruction": self.instruction,
                },
            )

        self.episode_step = 0
        self.last_reward = None
        self.last_step_result = None

        return self._make_observation(), info

    def step(self, action: Any) -> Tuple[Dict[str, Any], float, bool, bool, Dict[str, Any]]:
        if not self.env_id:
            raise RuntimeError("Environment not initialized. Call reset() first.")

        payload = dict(self.step_defaults)
        eval_payload_from_action = None
        evaluate_now = self.evaluate_every_step

        if isinstance(action, dict):
            action_map = dict(action)
            evaluate_now = bool(action_map.pop("evaluate", evaluate_now))
            eval_payload_from_action = action_map.pop("evaluate_payload", None)
            payload.update(action_map)
        else:
            payload["command"] = str(action)

        command = str(payload.get("command", "")).strip()
        if not command:
            raise ValueError("Action must provide a non-empty command")

        if self.interaction_logger is not None:
            self.interaction_logger.log_event(
                "sandbox.step.request",
                {
                    "env_id": self.env_id,
                    "step_index": self.episode_step + 1,
                    "payload": payload,
                },
            )

        step_result = self.client.step(self.env_id, payload)
        self.last_step_result = step_result
        self.episode_step += 1

        if self.interaction_logger is not None:
            self.interaction_logger.log_event(
                "sandbox.step.response",
                {
                    "env_id": self.env_id,
                    "step_index": self.episode_step,
                    "result": step_result,
                },
            )

        reward = 0.0
        terminated = False
        truncated = False
        eval_result = None

        if evaluate_now:
            eval_payload = dict(self.eval_defaults)
            if isinstance(eval_payload_from_action, dict):
                eval_payload.update(eval_payload_from_action)
            eval_result = self.client.evaluate(self.env_id, eval_payload)
            raw_reward = eval_result.get("reward")
            reward = float(raw_reward) if raw_reward is not None else 0.0
            self.last_reward = reward
            terminated = bool(eval_result.get("done", False))
            _print_and_log(
                "terminated={}, reward={}".format(terminated, reward),
                self.interaction_logger,
                event="sandbox.evaluate.response",
                payload={
                    "env_id": self.env_id,
                    "step_index": self.episode_step,
                    "result": eval_result,
                },
            )

        if self.episode_step >= self.max_episode_steps and not terminated:
            truncated = True

        info = {
            "env_id": self.env_id,
            "step_result": step_result,
            "evaluate_result": eval_result,
        }
        return self._make_observation(), reward, terminated, truncated, info

    def evaluate(self, payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        if not self.env_id:
            raise RuntimeError("Environment not initialized. Call reset() first.")
        req = dict(self.eval_defaults)
        if payload:
            req.update(payload)
        if self.interaction_logger is not None:
            self.interaction_logger.log_event(
                "sandbox.evaluate.request",
                {"env_id": self.env_id, "payload": req},
            )
        result = self.client.evaluate(self.env_id, req)
        raw_reward = result.get("reward")
        self.last_reward = float(raw_reward) if raw_reward is not None else None
        if self.interaction_logger is not None:
            self.interaction_logger.log_event(
                "sandbox.evaluate.response",
                {"env_id": self.env_id, "result": result},
            )
        return result

    def close(self) -> None:
        if self.env_id and self.auto_cleanup:
            try:
                self.client.delete_env(self.env_id, remove_image=True)
            except Exception:
                pass
        if self.interaction_logger is not None:
            self.interaction_logger.log_event(
                "sandbox.close",
                {"env_id": self.env_id, "auto_cleanup": self.auto_cleanup},
            )
        self.env_id = None


DONE_TOKEN = "<<DONE>>"


class BaseAgent(ABC):
    @abstractmethod
    def step(self, obs: str) -> str:
        """obs: terminal output -> action.

        Return DONE_TOKEN when the agent believes the task is complete.
        """
        raise NotImplementedError


class RuleBasedXlsxAgent(BaseAgent):
    """Deterministic demo agent for xlsx-recover-data task."""

    def __init__(self):
        self._used = False

    def step(self, obs: str) -> str:
        del obs
        if self._used:
            return "echo 'already solved'"
        self._used = True
        return textwrap.dedent(
            """
            python3 <<'PYCODE'
            import openpyxl
            wb = openpyxl.load_workbook("nasa_budget_incomplete.xlsx")
            budget = wb["Budget by Directorate"]
            yoy = wb["YoY Changes (%)"]
            shares = wb["Directorate Shares (%)"]
            growth = wb["Growth Analysis"]

            budget["F8"] = budget["K8"].value - sum(budget.cell(row=8, column=c).value for c in range(2, 11) if c != 6)
            budget["K5"] = sum(budget.cell(row=5, column=c).value for c in range(2, 11))
            yoy["D7"] = round((budget["D8"].value - budget["D7"].value) / budget["D7"].value * 100, 2)
            growth["B7"] = budget["B13"].value - budget["B8"].value
            budget["B9"] = round(budget["B8"].value * (1 + yoy["B8"].value / 100))
            budget["C12"] = round(budget["C11"].value * (1 + yoy["C11"].value / 100))
            budget["K10"] = round(budget["K9"].value * (1 + yoy["K9"].value / 100))
            yoy["F9"] = round((budget["F10"].value - budget["F9"].value) / budget["F9"].value * 100, 2)
            shares["F5"] = round(budget["F5"].value / budget["K5"].value * 100, 2)
            budget["E10"] = round(budget["K10"].value * shares["E10"].value / 100)
            yoy["B9"] = round((budget["B10"].value - budget["B9"].value) / budget["B9"].value * 100, 2)
            shares["B10"] = round(budget["B10"].value / budget["K10"].value * 100, 2)
            growth["B8"] = round((budget["B8"].value + budget["B9"].value + budget["B10"].value + budget["B11"].value + budget["B12"].value) / 5, 1)
            growth["E4"] = round(((budget["E13"].value / budget["E8"].value) ** 0.2 - 1) * 100, 2)
            growth["E5"] = budget["E8"].value

            wb.save("nasa_budget_recovered.xlsx")
            print("recovered")
            PYCODE
            """
        ).strip()


class OpenAICommandAgent(BaseAgent):
    """LLM command agent using OpenAI Responses API.

    Prompt architecture (per SkillsBench paper C.4/C.5):
    - System message: execution policy only.
    - First user message: task instruction + skills metadata + kickoff.
    - Later user messages: per-step execution feedback (stdout/stderr/exit_code).
      User messages are subject to sliding-window trimming (8K token budget).
    - Skills metadata (name, description, path) follows the OpenAI skills
      standard: agents discover skills autonomously from the metadata and
      read <path>/SKILL.md on demand via shell commands.
    """

    # Base system prompt contains policy only.
    BASE_SYSTEM_PROMPT = textwrap.dedent(
        """
        You are an autonomous command-line agent operating inside a SkillsBench sandbox container.
        Your objective is to solve the given task.

        Output contract:
        - Return exactly one shell command per turn.
        - When you believe the task is fully complete, output exactly: {done_token}
        - Do NOT output {done_token} until you are confident the task is solved.
        """
    ).strip().format(done_token=DONE_TOKEN)
    # BASE_SYSTEM_PROMPT = textwrap.dedent(
    #     """
    #     You are an autonomous command-line agent operating inside a SkillsBench sandbox container.
    #     Your objective is to solve the given task using the provided skills and maximize evaluator reward.

    #     Hard output contract:
    #     - Return exactly one shell command per turn.
    #     - No markdown, no explanation, no code fences.
    #     - Command must be executable in bash.

    #     Execution policy:
    #     - Follow the task instruction exactly, including file paths, names, schema, and units.
    #     - Prefer small, verifiable steps; inspect before editing when uncertain.
    #     - Preserve unrelated content and formatting in provided files.
    #     - Use the provided skills.
    #     """
    # ).strip()

    STEP_TEMPLATE = textwrap.dedent(
        """
        exit_code: {exit_code}
        stdout:
        {stdout}
        stderr:
        {stderr}
        """
    ).strip()

    def __init__(
        self,
        model: str,
        temperature: float = 0.0,
        max_output_tokens: int = 300,
        history_turns: Optional[int] = None,
        context_window_tokens: int = 160000,
        interaction_logger: Optional[InteractionLogger] = None,
    ):
        try:
            from openai import OpenAI  # type: ignore
        except ImportError:
            raise RuntimeError("Install OpenAI SDK first: pip install openai")

        self.client = OpenAI()
        self.model = model
        self.temperature = temperature
        self.max_output_tokens = max_output_tokens
        self.history_turns = None if history_turns is None else max(1, int(history_turns))
        self.context_window_tokens = max(1000, int(context_window_tokens))
        self.interaction_logger = interaction_logger
        self._env_id = None  # type: Optional[str]
        self._system_prompt = self.BASE_SYSTEM_PROMPT
        self._task_instruction = ""
        self._skills_section = ""
        self._has_pinned_first_user = False
        self._messages = []  # type: list
        self._reset_history()

    def set_task_context(self, instruction: str, skills_section: str) -> None:
        """Store task context for the first-turn user prompt."""
        self._task_instruction = instruction.strip()
        self._skills_section = skills_section.strip()
        self._reset_history()

    def _reset_history(self) -> None:
        self._has_pinned_first_user = False
        self._messages = [
            {"role": "system", "content": self._system_prompt},
        ]

    def _pinned_prefix_len(self) -> int:
        # Keep system prompt always; keep first user turn (task instruction + skills) once emitted.
        if self._has_pinned_first_user and len(self._messages) >= 2:
            return 2
        return 1

    def _build_first_turn_prompt(self) -> str:
        parts = []  # type: List[str]
        if self._task_instruction:
            parts.append("## Task Instruction\n\n" + self._task_instruction)
        if self._skills_section:
            parts.append(self._skills_section)
        parts.append("Begin. Return the first shell command.")
        return "\n\n".join(parts)

    def _trim_history(self) -> None:
        pinned = self._pinned_prefix_len()
        # Pinned: system message (+ first user turn after kickoff, if available).
        # Droppable: later user/assistant turns.
        # 1) Optional hard cap on droppable turns.
        if self.history_turns is not None:
            keep_tail = 2 * self.history_turns
            head = self._messages[:pinned]
            tail = self._messages[pinned:]
            if len(tail) > keep_tail:
                tail = tail[-keep_tail:]
            self._messages = head + tail
        # 2) Sliding window around target token budget (8K default per paper).
        while self._history_tokens() > self.context_window_tokens and len(self._messages) > pinned:
            del self._messages[pinned]

    def _history_tokens(self) -> int:
        total = 0
        for msg in self._messages:
            content = msg.get("content", "")
            total += max(1, len(str(content)) // 4)
        return total

    def _parse_obs(self, obs: str) -> Dict[str, Any]:
        try:
            data = json.loads(obs)
        except Exception:
            data = {}
        if isinstance(data, dict):
            return data
        return {}

    def _extract_text(self, response: Any) -> str:
        text = getattr(response, "output_text", "")
        if text:
            return text.strip()

        # Fallback for SDK versions without output_text convenience field.
        data = None
        try:
            data = response.model_dump()
        except Exception:
            data = None

        if isinstance(data, dict):
            chunks = []
            for item in data.get("output", []):
                for c in item.get("content", []):
                    maybe_text = c.get("text")
                    if maybe_text:
                        chunks.append(maybe_text)
            if chunks:
                return "\n".join(chunks).strip()

        return str(response).strip()

    def _extract_command(self, text: str) -> str:
        text = text.strip()
        if not text:
            return ""

        # Check if the LLM signals task completion.
        if DONE_TOKEN in text:
            return DONE_TOKEN

        fence = re.search(r"```(?:bash|sh)?\s*(.*?)```", text, flags=re.DOTALL | re.IGNORECASE)
        if fence:
            return fence.group(1).strip()

        text = re.sub(r"^\s*command\s*:\s*", "", text, flags=re.IGNORECASE)
        return text.strip()

    def step(self, obs: str) -> str:
        data = self._parse_obs(obs)
        env_id = str(data.get("env_id", ""))
        step_count = int(data.get("step_count", 0) or 0)
        if env_id != self._env_id:
            self._env_id = env_id
            self._reset_history()

        if step_count == 0:
            # First turn: send task instruction + skills as user input.
            self._reset_history()
            prompt = self._build_first_turn_prompt()
        else:
            # Subsequent turns: execution feedback only.
            prompt = self.STEP_TEMPLATE.format(
                exit_code=data.get("last_exit_code", ""),
                stdout=str(data.get("last_stdout", "")),
                stderr=str(data.get("last_stderr", "")),
            )
        _print_and_log(
            "step: {}, user: {}".format(step_count, prompt),
            self.interaction_logger,
            event="agent.prompt",
            payload={
                "env_id": env_id,
                "step_count": step_count,
                "prompt": prompt,
            },
        )
        self._messages.append({"role": "user", "content": prompt})
        if step_count == 0:
            self._has_pinned_first_user = True
        self._trim_history()

        response = self.client.responses.create(
            model=self.model,
            input=self._messages,
            temperature=self.temperature,
            # max_output_tokens=self.max_output_tokens,
            truncation="auto",
        )

        text = self._extract_text(response)
        command = self._extract_command(text)
        if not command:
            raise RuntimeError("LLM produced empty command")
        self._messages.append({"role": "assistant", "content": text})
        _print_and_log(
            "step: {}, assistant: {}".format(step_count, text),
            self.interaction_logger,
            event="agent.response",
            payload={
                "env_id": env_id,
                "step_count": step_count,
                "response": text,
                "command": command,
            },
        )
        self._trim_history()
        return command


def _format_skills_section(task_meta: Dict[str, Any], include_skill_metadata: bool) -> str:
    """Build the skills section for the first-turn user prompt (OpenAI skills style).

    Per OpenAI docs, skill instructions are user prompt input, not system prompt.
    We inject skill name/description/path so the model can decide which to load,
    then read <path>/SKILL.md inside the sandbox on demand.
    """
    skills = task_meta.get("skills_metadata", []) if include_skill_metadata else []
    if not skills:
        return ""

    lines = [
        "## Available Skills",
        "",
        "You have access to the skills listed below. Proactively use any relevant skill to complete the task accurately and efficiently.",
        "",
    ]
    for s in skills:
        lines.append("- **{}** (path: `{}`): {}".format(s["name"], s["path"], s["description"]))
    lines.append("\nTo use a relevant skill, you should read its full instructions: cat <path>/SKILL.md")
    return "\n".join(lines)


def run_rollout(
    env: SandboxGymEnv,
    agent: BaseAgent,
    max_steps: int,
    final_evaluate: bool,
    task_meta: Dict[str, Any],
    include_skill_metadata: bool,
    interaction_logger: Optional[InteractionLogger] = None,
) -> Dict[str, Any]:
    obs, info = env.reset()
    _print_and_log(
        "[reset] env_id={}".format(info.get("env_id")),
        interaction_logger,
        event="rollout.reset",
        payload={"info": info},
    )
    # print("[instruction]\n{}".format(obs.get("instruction", "").strip()))

    # Provide instruction + skills to agents that support explicit task context.
    # OpenAI agent places this context in first-turn user input.
    skills_section = _format_skills_section(task_meta, include_skill_metadata)
    if hasattr(agent, "set_task_context"):
        agent.set_task_context(
            instruction=str(obs.get("instruction", "")),
            skills_section=skills_section,
        )

    total_reward = 0.0
    terminated = False
    truncated = False
    agent_done = False

    for i in range(max_steps):
        agent_obs = {
            "env_id": str(info.get("env_id", "")),
            "step_count": int(obs.get("step_count", 0) or 0),
            "last_exit_code": obs.get("last_exit_code"),
            "last_stdout": str(obs.get("last_stdout", "")),
            "last_stderr": str(obs.get("last_stderr", "")),
        }
        if interaction_logger is not None:
            interaction_logger.log_event(
                "agent.observation",
                {"rollout_step": i + 1, "observation": agent_obs},
            )
        action = agent.step(json.dumps(agent_obs, ensure_ascii=True))
        if interaction_logger is not None:
            interaction_logger.log_event(
                "agent.action",
                {"rollout_step": i + 1, "action": action},
            )

        # Agent signals task completion — skip execution, go straight to evaluate.
        if action == DONE_TOKEN:
            _print_and_log(
                "[step {}] agent signaled {} - running evaluate".format(i + 1, DONE_TOKEN),
                interaction_logger,
                event="rollout.agent_done",
                payload={"rollout_step": i + 1},
            )
            agent_done = True
            break

        obs, reward, terminated, truncated, step_info = env.step(action)
        total_reward += reward

        step_result = step_info.get("step_result") or {}
        _print_and_log(
            "[step {}] exit_code={} duration_sec={} reward={}".format(
                i + 1,
                step_result.get("exit_code"),
                step_result.get("duration_sec"),
                reward,
            ),
            interaction_logger,
            event="rollout.step",
            payload={
                "rollout_step": i + 1,
                "step_result": step_result,
                "reward": reward,
            },
        )
        if terminated:
            _print_and_log(
                "[step {}] terminated=True, reward={}".format(i + 1, total_reward),
                interaction_logger,
                event="rollout.terminated",
                payload={"rollout_step": i + 1, "total_reward": total_reward},
            )
            break
        if truncated:
            _print_and_log(
                "[step {}] truncated=True, reward={}".format(i + 1, total_reward),
                interaction_logger,
                event="rollout.truncated",
                payload={"rollout_step": i + 1, "total_reward": total_reward},
            )
            break

    final_eval = None
    # Evaluate when the agent declares done, or as a final fallback if not terminated.
    if agent_done or (final_evaluate and not terminated):
        final_eval = env.evaluate()
        eval_reward = final_eval.get("reward")
        if eval_reward is not None:
            total_reward += float(eval_reward)
        terminated = bool(final_eval.get("done", False))
        _print_and_log(
            "\n[final evaluate] reward={} done={} exit_code={}".format(
                final_eval.get("reward"),
                final_eval.get("done"),
                final_eval.get("exit_code"),
            ),
            interaction_logger,
            event="rollout.final_evaluate",
            payload={"result": final_eval, "total_reward": total_reward},
        )

    result = {
        "terminated": terminated,
        "truncated": truncated,
        "total_reward": total_reward,
        "final_evaluate": final_eval,
        "env_id": env.env_id,
    }
    if interaction_logger is not None:
        interaction_logger.log_event("rollout.summary", result)
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Gym-style rollout example for SkillsBench sandbox")
    parser.add_argument("--server-url", default="http://127.0.0.1:8080")
    parser.add_argument(
        "--auto-start-server",
        action="store_true",
        help="Automatically start local sandbox server if /health is unreachable",
    )
    parser.add_argument(
        "--server-config",
        default=None,
        help="Config file path used when --auto-start-server (default: ./sandbox_config.yaml if exists)",
    )
    parser.add_argument(
        "--server-start-timeout-sec",
        type=int,
        default=20,
        help="Max seconds to wait for auto-started server health",
    )
    parser.add_argument(
        "--server-log-file",
        default="/tmp/skillsbench-sandbox-server.log",
        help="Log file for auto-started sandbox server",
    )

    task_group = parser.add_mutually_exclusive_group(required=True)
    task_group.add_argument("--task-id", help="Task id used with --dataset")
    task_group.add_argument("--task-path", help="Explicit task path, e.g. tasks/xlsx-recover-data")

    parser.add_argument("--dataset", default="tasks")
    parser.add_argument("--rebuild", action="store_true", help="Rebuild docker image when creating env")

    parser.add_argument("--agent", choices=["rule", "openai"], default="rule")
    parser.add_argument("--model", default="gpt-5.2-codex", help="OpenAI model when --agent=openai")
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument(
        "--history-turns",
        type=int,
        default=None,
        help="Optional hard cap on kept conversation turns (default: unlimited; token window still applies)",
    )
    parser.add_argument(
        "--context-window-tokens",
        type=int,
        default=8000,
        help="Sliding context window token budget (oldest turns dropped beyond this limit)",
    )
    parser.add_argument(
        "--disable-skill-injection",
        action="store_true",
        help="Disable OpenAI skills-style metadata injection in user observation context",
    )

    parser.add_argument("--max-steps", type=int, default=None, help="Default: inferred by difficulty (10/30/50)")
    parser.add_argument(
        "--max-episode-steps",
        type=int,
        default=None,
        help="Default: same as max-steps (difficulty-derived)",
    )
    parser.add_argument("--evaluate-every-step", action="store_true")
    parser.add_argument("--no-final-evaluate", action="store_true")

    parser.add_argument("--step-timeout-sec", type=int, default=None)
    parser.add_argument("--eval-timeout-sec", type=int, default=None)
    parser.add_argument("--workdir", default=None)
    parser.add_argument("--max-output-chars", type=int, default=None)
    parser.add_argument(
        "--interaction-log-file",
        default="./log.jsonl",
        help="Plain-text file that mirrors printed rollout logs",
    )

    parser.add_argument("--keep-env", action="store_true", help="Do not delete env on script exit")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    managed_server = None  # type: Optional[ManagedSandboxServer]
    interaction_logger = None  # type: Optional[InteractionLogger]
    project_root = Path(__file__).resolve().parents[1]
    resolved_task_path = _resolve_task_path(project_root, args.dataset, args.task_id, args.task_path)
    task_meta = _load_task_meta(resolved_task_path)

    interaction_log_path = None  # type: Optional[Path]
    if str(args.interaction_log_file or "").strip():
        interaction_log_path = Path(args.interaction_log_file).expanduser()
        interaction_logger = InteractionLogger(interaction_log_path)
        _print_and_log(
            "[interaction log] writing to {}".format(interaction_logger.path),
            interaction_logger,
            event="rollout.log_path",
            payload={"path": str(interaction_logger.path)},
        )

    default_rounds = int(task_meta.get("default_rounds", 30) or 30)
    max_steps = int(args.max_steps) if args.max_steps is not None else default_rounds
    max_episode_steps = int(args.max_episode_steps) if args.max_episode_steps is not None else max_steps

    step_timeout_sec = args.step_timeout_sec
    if step_timeout_sec is None and task_meta.get("agent_timeout_sec") is not None:
        step_timeout_sec = int(task_meta["agent_timeout_sec"])

    eval_timeout_sec = args.eval_timeout_sec
    if eval_timeout_sec is None and task_meta.get("verifier_timeout_sec") is not None:
        eval_timeout_sec = int(task_meta["verifier_timeout_sec"])

    if args.auto_start_server:
        managed_server = _start_server_if_needed(
            base_url=args.server_url,
            config_path=args.server_config,
            startup_timeout_sec=args.server_start_timeout_sec,
            log_path=Path(args.server_log_file).expanduser(),
            interaction_logger=interaction_logger,
        )
    elif not _is_server_healthy(args.server_url):
        raise RuntimeError(
            "Sandbox server is not reachable at {}. Start it first or use --auto-start-server.".format(
                args.server_url
            )
        )

    _print_and_log(
        "[task config] difficulty={} level={} max_rounds={} step_timeout_sec={} eval_timeout_sec={} skills={}".format(
            task_meta.get("difficulty", ""),
            task_meta.get("level", ""),
            max_steps,
            step_timeout_sec,
            eval_timeout_sec,
            len(task_meta.get("skills_metadata", []) or []),
        ),
        interaction_logger,
    )
    if interaction_logger is not None:
        interaction_logger.log_event(
            "rollout.task_config",
            {
                "difficulty": task_meta.get("difficulty", ""),
                "level": task_meta.get("level", ""),
                "max_rounds": max_steps,
                "step_timeout_sec": step_timeout_sec,
                "eval_timeout_sec": eval_timeout_sec,
                "skills_count": len(task_meta.get("skills_metadata", []) or []),
            },
        )

    step_defaults = {}
    if step_timeout_sec is not None:
        step_defaults["timeout_sec"] = step_timeout_sec
    if args.workdir:
        step_defaults["workdir"] = args.workdir
    if args.max_output_chars is not None:
        step_defaults["max_output_chars"] = args.max_output_chars

    eval_defaults = {}
    if eval_timeout_sec is not None:
        eval_defaults["timeout_sec"] = eval_timeout_sec
    if args.max_output_chars is not None:
        eval_defaults["max_output_chars"] = args.max_output_chars

    client = SandboxHTTPClient(args.server_url)
    env = SandboxGymEnv(
        client=client,
        dataset=args.dataset,
        task_id=args.task_id,
        task_path=args.task_path,
        rebuild=args.rebuild,
        max_episode_steps=max_episode_steps,
        evaluate_every_step=args.evaluate_every_step,
        step_defaults=step_defaults,
        eval_defaults=eval_defaults,
        auto_cleanup=not args.keep_env,
        interaction_logger=interaction_logger,
    )

    if args.agent == "openai":
        agent = OpenAICommandAgent(
            model=args.model,
            temperature=args.temperature,
            history_turns=args.history_turns,
            context_window_tokens=args.context_window_tokens,
            interaction_logger=interaction_logger,
        )
    else:
        agent = RuleBasedXlsxAgent()

    final_evaluate = (not args.no_final_evaluate) and (not args.evaluate_every_step)

    try:
        result = run_rollout(
            env,
            agent,
            max_steps=max_steps,
            final_evaluate=final_evaluate,
            task_meta=task_meta,
            include_skill_metadata=not args.disable_skill_injection,
            interaction_logger=interaction_logger,
        )
        _print_and_log(
            "\n[rollout summary] {}".format(json.dumps(result, ensure_ascii=True)),
            interaction_logger,
            event="rollout.summary.print",
            payload=result,
        )
    finally:
        env.close()
        if managed_server is not None:
            managed_server.stop()
            _print_and_log("[server] auto-started server stopped", interaction_logger)
        if interaction_logger is not None:
            interaction_logger.close()


if __name__ == "__main__":
    main()
